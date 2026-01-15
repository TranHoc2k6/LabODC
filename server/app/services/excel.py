import openpyxl
from pathlib import Path

class ExcelService:
    @staticmethod
    def load_task_template():
        template_path = Path("templates/task_template.xlsx")
        if not template_path.exists():
            return None
        return openpyxl.load_workbook(template_path)
    
    @staticmethod
    def parse_tasks_from_excel(file_path: str):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        tasks = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If task name exists
                tasks.append({
                    "title": row[0],
                    "description": row[1],
                    "assigned_to": row[2]
                })
        return tasks